#!/usr/bin/env python3
"""
run_scenarios_grid.py

Sweep many conversation-game scenarios, parse JSON output, and export a LARGE
Excel workbook with raw results, summaries, and embedded charts.

REQUIRES: pandas, openpyxl, matplotlib
DEFAULT RUNNER: "uv run" with entry "main.py"

EXAMPLES
--------
python run_scenarios_grid.py
python run_scenarios_grid.py --lengths 10 20 40 80 --subjects 10 20 --pr 2 4 6 8 --seeds 1 2
python run_scenarios_grid.py --runner "python" --entry main.py
python run_scenarios_grid.py --lengths 10 30 --pr 5 --seeds 42
python run_scenarios_grid.py --out results_sweep.xlsx
"""

import argparse
import itertools
import json
import os
import shlex
import subprocess
from dataclasses import dataclass
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


@dataclass
class Scenario:
	name: str
	subjects: int
	memory_size: int
	length: int
	seed: int
	p1_count: int
	pr_count: int


def build_command(runner: str, entry: str, sc: Scenario) -> list[str]:
	args = [
		runner,
		entry,
		'--subjects',
		str(sc.subjects),
		'--memory_size',
		str(sc.memory_size),
		'--length',
		str(sc.length),
		'--seed',
		str(sc.seed),
		'--player',
		'p1',
		str(sc.p1_count),
		'--player',
		'pr',
		str(sc.pr_count),
	]
	if ' ' in runner:
		return shlex.split(runner) + args[1:]
	return args


def run_and_capture(cmd: list[str]) -> str:
	proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
	if proc.returncode != 0:
		raise RuntimeError(
			f'Command failed ({proc.returncode}): {" ".join(cmd)}\nOutput:\n{proc.stdout}'
		)
	return proc.stdout


def extract_json_blob(stdout: str) -> dict[str, Any]:
	start = stdout.find('{')
	if start == -1:
		raise ValueError('Could not find JSON in output.')
	json_text = stdout[start:]
	try:
		return json.loads(json_text)
	except json.JSONDecodeError:
		end = json_text.rfind('}')
		if end != -1:
			return json.loads(json_text[: end + 1])
		raise


def find_player_id(payload: dict[str, Any], player_name: str = 'Player1') -> str | None:
	for turn in payload.get('turn_impact', []):
		if turn.get('speaker_name') == player_name:
			return turn.get('speaker_id')
	for p in payload.get('scores', {}).get('player_scores', []):
		if p.get('name') == player_name:
			return p.get('id')
	return None


def compute_quality(total: float, conv_len: int) -> float:
	return (total / conv_len) if conv_len else 0.0


def summarize(payload: dict[str, Any], player_name: str = 'Player1') -> dict[str, Any]:
	scores = payload.get('scores', {})
	conv_len = int(scores.get('conversation_length', 0))

	shared = payload.get('score_breakdown') or scores.get('shared_score_breakdown') or {}
	shared_total = float(shared.get('total', 0.0))
	shared_importance = float(shared.get('importance', 0.0))
	shared_coherence = float(shared.get('coherence', 0.0))
	shared_freshness = float(shared.get('freshness', 0.0))
	shared_nonmono = float(shared.get('nonmonotonousness', 0.0))

	p1_id = find_player_id(payload, player_name=player_name)
	p1_total = p1_shared = p1_individual = None
	p1_quality = None

	for p in scores.get('player_scores', []):
		if p1_id and p.get('id') == p1_id:
			p1_total = float(p['scores']['total'])
			p1_shared = float(p['scores']['shared'])
			p1_individual = float(p['scores']['individual'])
			p1_quality = compute_quality(p1_total, conv_len)
			break

	return {
		'conversation_length': conv_len,
		'shared_total': shared_total,
		'shared_importance': shared_importance,
		'shared_coherence': shared_coherence,
		'shared_freshness': shared_freshness,
		'shared_nonmonotonousness': shared_nonmono,
		'player1_id': p1_id,
		'player1_total': p1_total,
		'player1_shared': p1_shared,
		'player1_individual': p1_individual,
		'player1_quality': p1_quality,
	}


def make_charts(df: pd.DataFrame, out_png_dir: str, prefix: str = 'chart') -> list[str]:
	"""
	Create a small set of high-signal charts and save them to PNG files.
	Returns a list of PNG paths (in a stable order) for embedding in Excel.
	"""
	os.makedirs(out_png_dir, exist_ok=True)
	pngs: list[str] = []

	# 1) Avg Player1 quality by length (bar)
	g1 = df.groupby('length', dropna=False)['player1_quality'].mean()
	plt.figure(figsize=(6, 4))
	g1.plot(kind='bar')
	plt.title('Avg Player1 Quality by Length')
	plt.ylabel('Quality (Total / Length)')
	plt.xlabel('Length')
	p = os.path.join(out_png_dir, f'{prefix}_quality_by_length.png')
	plt.tight_layout()
	plt.savefig(p, dpi=160)
	plt.close()
	pngs.append(p)

	# 2) Avg Player1 quality by RandomPlayers (bar)
	g2 = df.groupby('pr_count', dropna=False)['player1_quality'].mean()
	plt.figure(figsize=(6, 4))
	g2.plot(kind='bar')
	plt.title('Avg Player1 Quality by RandomPlayers')
	plt.ylabel('Quality')
	plt.xlabel('RandomPlayers (pr_count)')
	p = os.path.join(out_png_dir, f'{prefix}_quality_by_pr.png')
	plt.tight_layout()
	plt.savefig(p, dpi=160)
	plt.close()
	pngs.append(p)

	# 3) Heatmap-like chart for Avg Quality by (length x pr_count)
	pivot = df.pivot_table(
		index='length', columns='pr_count', values='player1_quality', aggfunc='mean'
	)
	plt.figure(figsize=(6.5, 4.5))
	# Use imshow to avoid seaborn; handle NaNs by converting to 0 for display

	data = pivot.fillna(0).values
	plt.imshow(data, aspect='auto')
	plt.title('Avg Player1 Quality (Length x PR)')
	plt.colorbar(label='Quality')
	plt.yticks(range(len(pivot.index)), pivot.index)
	plt.xticks(range(len(pivot.columns)), pivot.columns)
	plt.xlabel('RandomPlayers (pr_count)')
	plt.ylabel('Length')
	p = os.path.join(out_png_dir, f'{prefix}_quality_heatmap_len_pr.png')
	plt.tight_layout()
	plt.savefig(p, dpi=160)
	plt.close()
	pngs.append(p)

	# 4) Shared breakdown by length (stacked bars)
	g3 = df.groupby('length', dropna=False)[
		['shared_importance', 'shared_coherence', 'shared_freshness', 'shared_nonmonotonousness']
	].mean()
	plt.figure(figsize=(7.5, 4.5))
	bottoms = None
	for col in [
		'shared_importance',
		'shared_coherence',
		'shared_freshness',
		'shared_nonmonotonousness',
	]:
		if bottoms is None:
			plt.bar(g3.index.astype(str), g3[col], label=col)
			bottoms = g3[col].copy()
		else:
			plt.bar(g3.index.astype(str), g3[col], bottom=bottoms, label=col)
			bottoms = bottoms + g3[col]
	plt.title('Shared Breakdown by Length (avg)')
	plt.xlabel('Length')
	plt.ylabel('Score')
	plt.legend(loc='best', fontsize=8)
	p = os.path.join(out_png_dir, f'{prefix}_shared_breakdown_by_length.png')
	plt.tight_layout()
	plt.savefig(p, dpi=160)
	plt.close()
	pngs.append(p)

	# 5) Avg Player1 quality by Memory Size
	g4 = df.groupby('memory_size', dropna=False)['player1_quality'].mean()
	plt.figure(figsize=(6, 4))
	g4.plot(kind='bar', color='teal')
	plt.title('Avg Player1 Quality by Memory Size')
	plt.ylabel('Quality (Total / Length)')
	plt.xlabel('Memory Size')
	p = os.path.join(out_png_dir, f'{prefix}_quality_by_memory_size.png')
	plt.tight_layout()
	plt.savefig(p, dpi=160)
	plt.close()
	pngs.append(p)

	return pngs


def embed_images_in_excel(
	xlsx_path: str, image_paths: list[str], sheet_name: str = 'charts'
) -> None:
	"""
	Create (or overwrite) a `charts` sheet and paste PNGs vertically.
	"""
	wb = load_workbook(xlsx_path)
	if sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		wb.remove(ws)
	ws = wb.create_sheet(sheet_name)

	row = 1
	for p in image_paths:
		if not os.path.exists(p):
			continue
		img = XLImage(p)
		# Anchor images roughly every ~22 rows to avoid overlap
		cell = f'A{row}'
		ws.add_image(img, cell)
		row += 22

	wb.save(xlsx_path)


def main():
	ap = argparse.ArgumentParser()
	ap.add_argument('--runner', default='uv run', help="Command to run (default: 'uv run')")
	ap.add_argument('--entry', default='main.py', help='Entry point (default: main.py)')
	ap.add_argument(
		'--subjects', type=int, nargs='+', default=[20], help='List of subject counts to sweep'
	)
	ap.add_argument(
		'--memory_sizes',
		type=int,
		nargs='+',
		default=[10, 20, 30],
		help='List of memory sizes to sweep',
	)
	ap.add_argument(
		'--lengths',
		type=int,
		nargs='+',
		default=[10, 30, 60],
		help='List of conversation lengths to sweep',
	)
	ap.add_argument(
		'--pr',
		type=int,
		nargs='+',
		default=[3, 5, 7, 10],
		help='List of RandomPlayer counts per scenario',
	)
	ap.add_argument('--p1', type=int, default=1, help='Number of Player1 instances (default 1)')
	ap.add_argument(
		'--seeds',
		type=int,
		nargs='+',
		default=[13],
		# default=[random.randint(1, 1000) for _ in range(3)],
		help='List of seeds to sweep (default: 3 random numbers)',
	)
	ap.add_argument(
		'--player_name', default='Player1', help='Name to identify your player in output'
	)
	ap.add_argument('--out', default='results_sweep.xlsx', help='Path to Excel workbook to write')
	ap.add_argument(
		'--charts_dir',
		default=None,
		help='Directory to write PNG charts (default: alongside Excel)',
	)
	ap.add_argument('--no_charts', action='store_true', help='Disable chart generation/embedding')
	args = ap.parse_args()

	all_rows = []
	idx = 0

	for idx, (S, B, L, PR, seed) in enumerate(
		itertools.product(args.subjects, args.memory_sizes, args.lengths, args.pr, args.seeds),
		start=1,
	):
		name = f'S{S}_B{B}_L{L}_PR{PR}_seed{seed}'
		sc = Scenario(
			name=name, subjects=S, memory_size=B, length=L, seed=seed, p1_count=args.p1, pr_count=PR
		)
		cmd = build_command(args.runner, args.entry, sc)
		print(f'\n[{idx}] Running {name}')
		print('Command:', ' '.join(shlex.quote(c) for c in cmd))

		try:
			stdout = run_and_capture(cmd)
			payload = extract_json_blob(stdout)
			summary = summarize(payload, player_name=args.player_name)
		except Exception as e:
			summary = {
				'conversation_length': None,
				'shared_total': None,
				'shared_importance': None,
				'shared_coherence': None,
				'shared_freshness': None,
				'shared_nonmonotonousness': None,
				'player1_id': None,
				'player1_total': None,
				'player1_shared': None,
				'player1_individual': None,
				'player1_quality': None,
				'error': str(e),
			}

		row = {
			'scenario': sc.name,
			'subjects': sc.subjects,
			'memory_size': sc.memory_size,
			'length': sc.length,
			'seed': sc.seed,
			'p1_count': sc.p1_count,
			'pr_count': sc.pr_count,
			**summary,
		}
		all_rows.append(row)

	# Build DataFrame
	df = pd.DataFrame(all_rows)

	# Create simple summaries
	grp1 = df.groupby(['length', 'pr_count'], dropna=False)['player1_quality'].mean().reset_index()
	grp1.rename(columns={'player1_quality': 'avg_player1_quality'}, inplace=True)

	grp2 = (
		df.groupby(['length', 'pr_count'], dropna=False)[
			[
				'shared_total',
				'shared_importance',
				'shared_coherence',
				'shared_freshness',
				'shared_nonmonotonousness',
			]
		]
		.mean()
		.reset_index()
	)

	grp3 = (
		df.groupby(['subjects', 'memory_size', 'length'], dropna=False)[
			['player1_total', 'player1_shared', 'player1_individual', 'player1_quality']
		]
		.mean()
		.reset_index()
	)

	# Write Excel
	out_xlsx = args.out
	with pd.ExcelWriter(out_xlsx, engine='openpyxl') as xlw:
		df.to_excel(xlw, index=False, sheet_name='all_results')
		grp1.to_excel(xlw, index=False, sheet_name='quality_by_len_PR')
		grp2.to_excel(xlw, index=False, sheet_name='shared_by_len_PR')
		grp3.to_excel(xlw, index=False, sheet_name='avgs_by_S_B_L')

	print(f'\nWrote Excel workbook to {out_xlsx}')
	print(f'Rows: {len(df)} | Unique scenarios: {df["scenario"].nunique()}')

	# Charts (PNG) + embed into Excel
	if not args.no_charts:
		png_dir = args.charts_dir or os.path.splitext(out_xlsx)[0] + '_charts'
		images = make_charts(df, png_dir, prefix='grid')
		try:
			embed_images_in_excel(out_xlsx, images, sheet_name='charts')
			print(f"Embedded {len(images)} charts into '{out_xlsx}' (sheet: charts)")
			print(f'PNG copies saved to: {png_dir}')
		except Exception as e:
			print(f'Could not embed images into Excel: {e}')
			print(f'Charts are still saved at: {png_dir}')


if __name__ == '__main__':
	main()
